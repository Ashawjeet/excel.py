from typing import List, Dict, Union
from openpyxl import load_workbook

# Load the Excel workbook
wb = load_workbook(filename="C:/Users/ashu shinde/Downloads/Sample Data Format (3).xlsx")

# Initialize dictionaries to store extracted data
data = {
    "Jobs": [],
    "Machine": [],
    "Priority": {},
    "Penalty": {},
    "Desired Endtime": {},
    "Operation": {},
    "Sequence": {},
    "Number of Jobs": {},
    "Job Details": {},
    "Available Job": {},
    "Available Machine": {},
    "Penalty Imposed": {},
    "Machine Downtime": {},
    "Plant Downtime": [],
    "Operator Downtime": {},
    "Operator": {},
    "Available Operator": {},
    "Fixture": {},
    "Available Fixture": {},
    "Result": [],
    "Total Time": {}
}

# Iterate over each sheet in the workbook
for sheet in wb:
    # Iterate over rows starting from row 2 (assuming headers are in row 1)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip rows that don't have enough columns
        if len(row) < 21:
            continue

        # Extract data and populate dictionaries
        job_id, machine_id, job_priority, job_penalty, job_desired_endtime, operation_details, sequence_details, num_of_jobs, job_specific_details, available_job_time, available_machine_time, penalty_cost, machine_downtime_details, plant_downtime_info, operator_downtime_details, operator_id, available_operator_time, fixture_details, available_fixture_time, result_data, total_time_details = row

        # Assign extracted values to variables
        data["Jobs"].append(job_id)
        data["Machine"].append(machine_id)
        data["Priority"][job_id] = job_priority
        data["Penalty"][job_id] = job_penalty
        data["Desired Endtime"][job_id] = job_desired_endtime
        data["Operation"][job_id] = operation_details
        data["Sequence"][job_id] = sequence_details
        data["Number of Jobs"][job_id] = num_of_jobs
        data["Job Details"][job_id] = job_specific_details
        data["Available Job"][job_id] = available_job_time
        data["Available Machine"][machine_id] = available_machine_time
        data["Penalty Imposed"][job_id] = penalty_cost
        data["Machine Downtime"][machine_id] = machine_downtime_details
        data["Plant Downtime"].append(plant_downtime_info)
        data["Operator Downtime"][operator_id] = operator_downtime_details
        data["Operator"][machine_id] = operator_id
        data["Available Operator"][operator_id] = available_operator_time
        data["Fixture"][machine_id] = fixture_details
        data["Available Fixture"][fixture_details] = available_fixture_time
        data["Result"].append(result_data)
        data["Total Time"][job_id] = total_time_details

# Assign extracted data to the previously defined variables
jobs: List[int] = data["Jobs"]
machine: List[int] = data["Machine"]
priority: Dict[int, int] = data["Priority"]
penalty: Dict[int, bool] = data["Penalty"]
desired_endtime: Dict[int, str] = data["Desired Endtime"]
operation: Dict[int, Dict[int, int]] = data["Operation"]
sequence: Dict[int, Union[List[int], List[List[int]]]] = data["Sequence"]
no_of_jobs: Dict[int, int] = data["Number of Jobs"]
job_details: Dict[int, Dict[str, Union[int, str, bool]]] = data["Job Details"]
available_job: Dict[int, str] = data["Available Job"]
available_machine: Dict[int, str] = data["Available Machine"]
penalty_imposed: Dict[int, int] = data["Penalty Imposed"]
machine_downtime: Dict[int, Dict[str, str]] = data["Machine Downtime"]
plant_downtime: List[Dict[str, Union[str, List[str]]]] = data["Plant Downtime"]
operator_downtime: Dict[int, Dict[str, str]] = data["Operator Downtime"]
operator: Dict[int, Union[int, List[int]]] = data["Operator"]
available_operator: Dict[int, str] = data["Available Operator"]
fixture: Dict[int, Union[int, List[int]]] = data["Fixture"]
available_fixture: Dict[int, str] = data["Available Fixture"]
result: List = data["Result"]
totalTime: Dict[int, Dict[int, int]] = data["Total Time"]
